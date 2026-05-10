import json
from datetime import datetime
from agents.base_agent import BaseAgent
from config import PHASE2_PROVIDER, PHASE2_MODEL, REPORTS_DIR


def _write_excel(ticker: str, run_id: int, raw: dict) -> str:
    """Write a simple Excel model from the raw_output dict. Returns file path string."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DCF Summary"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")

    dcf = raw.get("dcf", {})
    headers = ["Scenario", "Intrinsic Value", "Discount Rate", "Terminal Growth"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, scenario in enumerate(["base", "bull", "bear"], 2):
        ws.cell(row=row_idx, column=1, value=scenario.capitalize())
        ws.cell(row=row_idx, column=2, value=dcf.get(scenario))
        ws.cell(row=row_idx, column=3, value=dcf.get("discount_rate"))
        ws.cell(row=row_idx, column=4, value=dcf.get("terminal_growth"))

    ws2 = wb.create_sheet("Expected Returns")
    horizons = ["6m", "1y", "2y", "5y", "10y"]
    ws2.cell(row=1, column=1, value="Horizon")
    for col, scenario in enumerate(["Bear", "Base", "Bull"], 2):
        cell = ws2.cell(row=1, column=col, value=scenario)
        cell.fill = header_fill
        cell.font = header_font

    returns = raw.get("expected_returns", {})
    for row_idx, horizon in enumerate(horizons, 2):
        ws2.cell(row=row_idx, column=1, value=horizon)
        h_data = returns.get(horizon, {})
        ws2.cell(row=row_idx, column=2, value=h_data.get("bear"))
        ws2.cell(row=row_idx, column=3, value=h_data.get("base"))
        ws2.cell(row=row_idx, column=4, value=h_data.get("bull"))

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    date_str = datetime.utcnow().strftime("%Y%m%d")
    file_name = f"{ticker}_{date_str}_model.xlsx"
    file_path = REPORTS_DIR / file_name
    wb.save(str(file_path))
    return str(file_path)


class FinancialModelerAgent(BaseAgent):
    name = "financial_modeler"
    phase = 2
    role_file = "financial_modeler.md"
    skill_files = ["dcf_valuation.md", "financial_ratio_analysis.md"]
    provider = PHASE2_PROVIDER
    model = PHASE2_MODEL

    def _build_user_prompt(self, bundle: dict) -> str:
        data = bundle.get("data", {})
        manifest = bundle.get("manifest", {})
        relevant = {
            "ticker": bundle.get("ticker"),
            "data_manifest": manifest,
            "income_statement": data.get("income_statement"),
            "balance_sheet": data.get("balance_sheet"),
            "cash_flow": data.get("cash_flow"),
            "fundamentals": data.get("fundamentals"),
            "earnings": data.get("earnings"),
            "live_price": data.get("live_price"),
            "phase1_summaries": bundle.get("phase1_summaries", {}),
        }
        return f"Build a financial model and DCF valuation for this stock and respond with the required JSON:\n\n{json.dumps(relevant, default=str)}"

    def run(self, bundle: dict, run_id: int) -> dict:
        result = super().run(bundle, run_id)
        if result["status"] == "complete":
            ticker = bundle.get("ticker", "UNKNOWN")
            try:
                model_path = _write_excel(ticker, run_id, result["raw_output"])
                result["raw_output"]["model_path"] = model_path
            except Exception as e:
                result["raw_output"]["model_path"] = None
                result["raw_output"]["excel_error"] = str(e)
        return result
